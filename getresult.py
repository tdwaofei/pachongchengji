import openpyxl
import requests
from PIL import Image
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import base64
from huaweicloudsdkcore.auth.credentials import BasicCredentials
from huaweicloudsdkocr.v1.region.ocr_region import OcrRegion
from huaweicloudsdkcore.exceptions import exceptions
from huaweicloudsdkocr.v1 import *
import ocrimage
import io

# 加载Excel表格
wb = openpyxl.load_workbook('D:\\gaokaostudent1.xlsx')
sheet = wb.active

# 创建WebDriver对象
driver = webdriver.Chrome()

# 打开查分网站
headers = {"user-agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/114.0'}
driver.get('http://query1.bjeea.cn/queryService/rest/score/870810')  # 替换为实际的查分网站链接
# 等待页面加载完成，直到指定元素可见
wait = WebDriverWait(driver, 10)  # 设置最大等待时间为10秒
#element = wait.until(EC.visibility_of_element_located((By.XPATH, "/html/body/app-root/app-main/div/div/div/app-score-query/div/div/app-score-query-form/div/div[1]/div/div/div[2]/form/div[1]/div[2]/tui-input/tui-hosted-dropdown/div/tui-primitive-textfield/tui-wrapper/div/div/div/tui-value-decoration/span[2]")))
#element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "app-score-query-form tui-value-decoration span:nth-child(2)")))
#print(element)
# 循环读取Excel表格中的考生号和身份证
for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    examNo_id, examinneNo_id,idcard_id = row[0], row[1],row[4]
    #截取idcard_id的后6位
    idcard6_id = idcard_id[-6:]
    # 输入考生号和身份证
    examNo_input = driver.find_element(By.ID, 'examNo')  # 准考证号
    examinneNo_input = driver.find_element(By.ID, 'examinneNo')  # 考生号
    idcard_input = driver.find_element(By.ID, 'idCard')#idCard身份证号
    captcha_input = driver.find_element(By.ID, 'captcha')  # 验证码
    #准考证号
    examNo_input.clear()
    examNo_input.send_keys(examNo_id)
    #考生号
    examinneNo_input.clear()
    examinneNo_input.send_keys(examinneNo_id)
    #身份证后6位
    idcard_input.clear()
    idcard_input.send_keys(idcard6_id)
    #获取验证码下载到本地
    captcha_input_img = driver.find_element(By.ID, 'captchaImg')  # 验证码
    captcha_input_img.screenshot('D:\captcha11.png')
    # 优化图片
    def preprocess_image(image):
        # 灰度化
        image = image.convert('L')
        # 二值化
        threshold = 152  # 二值化阈值
        image = image.point(lambda x: 0 if x < threshold else 255, '1')
        return image
    # 把保存在本地的验证码图片转换为jpg格式
    image = Image.open('D:\captcha11.png')
    preprocess_captcha_image_jpg = preprocess_image(image)
    image = preprocess_captcha_image_jpg.convert('RGB')
    image.save('D:\processed_captcha11.jpg')
    #调用华为OCR识别验证码
    ak = "8TZ6BMUOMWO9BORZI0GG"
    sk = "2bFlDAwKF7zidUYHOzUfnoTuI3hKDOwyp7ClLgAt"
    credentials = BasicCredentials(ak, sk)  # 构造鉴权对象
    client = OcrClient.new_builder() \
        .with_credentials(credentials) \
        .with_region(OcrRegion.value_of("cn-north-4")) \
        .build()

    # 读取图片转64位编码
    def encode_image(origin_image):
        with io.BytesIO() as buffer:
            origin_image.save(buffer, format="JPEG")
            return base64.b64encode(buffer.getvalue()).decode()
    # 识别图片
    try:
        request = RecognizeGeneralTextRequest()
        #从本地读取图片
        captcha_image_jpg = Image.open('D:\processed_captcha11.jpg')
        image64=encode_image(captcha_image_jpg)
        request.body = GeneralTextRequestBody(image=image64)
        response = client.recognize_general_text(request)
        response_dict = response.to_dict()
        words_block_list = response_dict.get('result', {}).get('words_block_list', [])
        if words_block_list:
            captcha_text = words_block_list[0]['words']
            print(captcha_text)
        else:
            print("No text found in image.")
    except exceptions.ClientRequestException as e:
        print(e.status_code)
        print(e.request_id)
        print(e.error_code)
        print(e.error_msg)

    # 输入验证码
    captcha_input.clear()
    captcha_input.send_keys(captcha_text)

    # 提交表单
    #submit_button = driver.find_element_by_xpath('//input[@id="queryBtn"]') # 替换为实际的提交按钮元素定位方式
    submit_button = driver.find_element(By.ID, 'queryBtn')  # 替换为实际的提交按钮元素定位方式
    submit_button.click()

    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'result_content')))  # 替换为实际的成绩表格元素定位方式
    # 解析成绩页面并获取成绩信息

    soup = BeautifulSoup(driver.page_source, 'lxml')  # 使用更快的解析器lxml
    score_table = soup.find('table', {'class': 'case'})  # 使用字典形式的参数传递，定位方式更清晰
    # 提取各科目成绩，根据实际网页结构提取相应的数据
    scores = score_table.find_all('td', {'class': 'title_p'})  # 使用find_all()函数，避免重复定位
    #获取scores标签信息中的语文、数学、英语成绩
    chinese_score = scores[0].find_next_sibling('td').text.strip() #语文成绩
    math_score = scores[2].find_next_sibling('td').text.strip() #数学成绩
    english_score = scores[4].find_next_sibling('td').text.strip()#英语成绩
    physics_score = scores[6].find_next_sibling('td').text.strip()#物理成绩
    history_score = scores[8].find_next_sibling('td').text.strip()#历史成绩
    geography_score = scores[10].find_next_sibling('td').text.strip()#地理成绩
    total_score = scores[12].find_next_sibling('td').text.strip()#总成绩
    # 将成绩写入Excel表格

    sheet.cell(row=row_number, column=7, value=chinese_score)
    sheet.cell(row=row_number, column=8, value=math_score)
    sheet.cell(row=row_number, column=9, value=english_score)
    sheet.cell(row=row_number, column=10, value=physics_score)
    sheet.cell(row=row_number, column=11, value=history_score)
    sheet.cell(row=row_number, column=12, value=geography_score)
    sheet.cell(row=row_number, column=13, value=total_score)

    # 点击返回按钮，返回到查询页面
    submit_button = driver.find_element(By.ID, 'backBtn')  # 替换为实际的提交按钮元素定位方式
    submit_button.click()
# 保存Excel表格
wb.save('D:\scores.xlsx')

# 关闭WebDriver
driver.quit()